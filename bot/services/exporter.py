"""Экспорт данных в Excel."""

from datetime import datetime
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from bot.config import config
from bot.services.parser import ParsedUser


class ExcelExporter:
    """Экспортер данных в Excel."""
    
    @staticmethod
    def export(users: List[ParsedUser], chat_title: str) -> Path:
        """
        Экспорт пользователей в Excel файл.
        
        Args:
            users: список пользователей
            chat_title: название чата (для имени файла)
        
        Returns:
            Путь к созданному файлу
        """
        # Конвертируем в DataFrame
        data = [user.to_dict() for user in users]
        df = pd.DataFrame(data)
        
        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = "".join(c if c.isalnum() or c in ' -_' else '_' for c in str(chat_title))
        filename = f"parsing_{safe_title}_{timestamp}.xlsx"
        filepath = config.exports_dir / filename
        
        # Сохраняем в Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Применяем форматирование
        ExcelExporter._apply_formatting(filepath)
        
        return filepath
    
    @staticmethod
    def _apply_formatting(filepath: Path):
        """Применение форматирования к Excel файлу."""
        from openpyxl import load_workbook
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Форматируем заголовки
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Автофильтр
        ws.auto_filter.ref = ws.dimensions
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)


# Глобальный экземпляр экспортера
exporter = ExcelExporter()